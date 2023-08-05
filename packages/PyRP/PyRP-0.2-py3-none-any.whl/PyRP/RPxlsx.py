import openpyxl as xlsx
from openpyxl.cell import get_column_letter
from openpyxl.styles import Font

if not __package__:
    from RP import Tournament as csvTournament, TournamentError
else:
    from .RP import Tournament as csvTournament, TournamentError


class Tournament(csvTournament):
    '''
    Reimplement the Tournament class so that it writes results to an xlsx file
    instead of a csv file.
    '''
    def run(
        self, methods_list, output_file=None, sort_by=None,
        reverse_order=True, heading=True, **kwargs):
        # Reimplements the parent version to write on an xlsx file
        wb = xlsx.Workbook()
        sheet = wb.active
        sheet.title = 'RP Tie-breaks'
        if heading:
            first_row = 2
            titles = (self._str_heading(**method) for method in methods_list)
            font = Font(bold=True)
            for column, title in enumerate(titles, 1):
                cell = sheet.cell(row=1, column=column)
                cell.value = title
                cell.font = font
        else:
            first_row = 1
        methods = [method['method'] for method in methods_list]
        data = self.results_list(methods_list, sort_by, reverse_order)
        for row, line in enumerate(data, first_row):
            for column, (value, method) in enumerate(zip(line, methods), 1):
                cell = sheet.cell(row=row, column=column)
                cell.value = value
                self._cell_style(cell, method)
        self._names_width(sheet, methods)
        try:
            wb.save(output_file)
        except OSError as error:
            text = self._FILE_ERROR.format(str(output_file))
            raise TournamentError(text) from error

    @staticmethod
    def _names_width(sheet, methods):
        try:
            index = methods.index('Names') + 1
        except ValueError:
            pass
        else:
            sheet.column_dimensions[get_column_letter(index)].width = 25

    _NAMES_FONT = Font(italic=True)
    _INTEGER = '0'
    _ONE_DECIMAL = '0.0'
    _THREE_DECIMAL = '0.000'

    @classmethod
    def _cell_style(cls, cell, method):
        if method == 'Names':
            cell.font = cls._NAMES_FONT
        elif method == 'Elos':
            cell.number_format = cls._INTEGER
        elif method in ('Points', 'Played points', 'Bucholz'):
            cell.number_format = cls._ONE_DECIMAL
        else:
            cell.number_format = cls._THREE_DECIMAL
    

###############################################################################
if __name__ == '__main__':
    from os.path import join
    t = Tournament.load(join('..', 'test', 'testing.txt'))
    t.run(
        methods_list=(
            {'method': 'Names'}, {'method': 'Points'},
            {'method': 'Elos'}, {'method': 'Average Bucholz'},
            {'method': 'Performance'}, {'method': 'RB'}, {'method': 'RP'},
            {'method': 'ARPO', 'worst': 1},
            {'method': 'ARPO', 'worst': 2},
            {'method': 'ARPO', 'worst': 1, 'best': 1}),
        output_file=join('..', 'test', 'testing.xlsx'),
        sort_by=(
            {'method': 'Points'},
            {'method': 'ARPO', 'worst': 1, 'best': 1}))
