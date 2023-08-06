import openpyxl, re, sys, pickle, pprint
from openpyxl.utils import column_index_from_string as cis

# created by hyunjin jeong, all right reserved

# open excel file / return openpyxl.workbook class
def open_excel_file(excel_file, read_only=True):
	try:
		exFile = openpyxl.load_workbook(excel_file,
					read_only=read_only,data_only=True)
		return exFile 
	except IOError as err:
		print("ioerror: "+str(err))
		sys.exit()

# id_generator
def id_gen(int_number, id_prefix_str, rjust_number=5):
	strCounter = str(int_number).rjust(rjust_number,"0")
	if id_prefix_str != None:
		strCounter = id_prefix_str+strCounter
	return strCounter

# return netestd dictionary 
#			{'key':{level2_key:value, ...},'key1':{leve2_key1:vlaue, ...}}
# 		worksheet: openpyxl.worksheet class
#		row_number_key: 키로 쓰일 행번호
#		column_number_key: 키로 쓰일 열번호 (default: None)
#		startCellIndex: 실제 값의 시작셀 인덱스 (ex. 'N27')
#		endCellIndex: 실제 값의 최종셀 인덱스 (ex. 'Z99')
#		id_prefix: column 키값이 없을 때 카운터로 대신한다.
#					이때 고유번호 id의 접두사를 결정 (default: None)
#		space_in_kw: 키워드 속에 공백문자 허용을 결정
def get_dictionary(
		worksheet, row_number_key, startCellIndex, endCellIndex,
		column_number_keys=None, id_prefix=None, space_in_kword=True
		):
	tmpDict = {}
	regexCellIndex = re.compile(r"^([A-Z]+)(\d+)$")
	sCell=regexCellIndex.search(startCellIndex)
	eCell=regexCellIndex.search(endCellIndex)

	if not(
			isinstance(sCell.group(0),str) 
			and isinstance(eCell.group(0),str)
			):
		print("invalid cell index")
		sys.exit()
	else:
		sColNo=cis(sCell.group(1))
		eColNo=cis(eCell.group(1))
		sRowNo=int(sCell.group(2))
		eRowNo=int(eCell.group(2))

	if space_in_kword == False:
		row_keys=[
			str(
				worksheet.cell(row=row_number_key, column=i).value
				).strip().replace(' ','_')
			for i in range(sColNo, eColNo+1)
			] # 키워드 안에 공백문자를 불허할 때 결과
	else:
		row_keys=[
			worksheet.cell(row=row_number_key, column=i).value 
			for i in range(sColNo, eColNo+1)
			] # list comprehension: 특정 행을 key로 지정
	
	table=worksheet[startCellIndex:endCellIndex]

	if column_number_keys != None:
		column_keys=[
			worksheet.cell(row=i, column=column_number_keys).value 
			for i in range(sRowNo, eRowNo+1)
			] # list comprehension: 특정 열을 key로 지정
		k=0
		for row in table:
			tmpDC={}
			for j in range(len(row_keys)):
				tmpDC[row_keys[j]]=row[j].value
			tmpDict[column_keys[k]]=tmpDC
			k += 1
	else:
		i=0
		for row in table:
			tmpDC={}
			for j in range(len(row_keys)):
				tmpDC[row_keys[j]]=row[j].value
			tmpDict[id_gen(i+1, id_prefix)]=tmpDC 
			i += 1

	return tmpDict

# save dictionary to pickle file
def save_dictionary_to_file(data_dict, save_filename):
	try:
		f = open(save_filename,'wb')
		pickle.dump(data_dict, f)
	except IOError as err:
		print("file error: "+str(err))
		sys.exit()
	except pickle.PickleError as pErr:
		print("pickle error: "+str(pErr))
		sys.exit()

# load dictionary from pickle file
def load_dictionary_from_file(load_filename):
	try:
		f = open(load_filename,'rb')
		fileDict = pickle.load(f)
	except IOError as err:
		print("file error: "+str(err))
		sys.exit()
	except pickle.PickleError as pErr:
		print("pickle error: "+str(pErr))
		sys.exit()

	return fileDict

# print the nested dictionary
# nested dictionary:
#		{'key':{level2_key:value, ...},'key1':{leve2_key1:vlaue, ...}}
def print_dictionary(nested_dictionary, seperatator=None):
	for key in sorted(nested_dictionary.keys()):
		print(str(key)+":")
		pprint.pprint(nested_dictionary[key])
		if seperatator != None:
			print(str(seperatator)*45)

# get highest row
def get_max_row(worksheet):
	return worksheet.max_row
