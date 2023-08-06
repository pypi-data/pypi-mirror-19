from collections import OrderedDict
from itertools import takewhile, groupby

from openpyxl import load_workbook
from yaml_dump import yaml_dump

from values_rows_to_js_nested_headers import values_rows_to_js_nested_headers


class Item(object):
    def __init__(self, sheet, title, slug, name_cells_rows):
        super().__init__()
        self.slug = slug
        self.sheet = sheet
        self.title = title
        self.header_cells_rows = name_cells_rows.get('header', [])
        self.names_cells = name_cells_rows.get('name')[0]
        self.names = [i.value for i in self.names_cells]

        types = name_cells_rows.get('type', [None] * len(self.names_cells))[0]

        self.types = [c.value for c in types]

        def get_sample_by_type(type, sample):
            v = sample.value
            if type == 'text':
                if not v:
                    return ''
                return str(v)
            if type == 'bool':
                return bool(v)
            if type == 'float':
                return float(v)
            return v

        samples = [[get_sample_by_type(type, sample) for type, sample in zip(self.types, i)] for i in
                   name_cells_rows.get('sample', [])]
        self.samples = [dict(zip(self.names, sample)) for sample in samples]

    @property
    def nested_headers(self):
        values_rows = [[cell.value for cell in cells_row] for cells_row in self.header_cells_rows]
        return values_rows_to_js_nested_headers(values_rows)

    @property
    def columns(self):
        l = [dict(data=name, type=type, width=width) for name, type, width in zip(self.names, self.types, self.widths)]

        def map_type(i):
            if i.get('type') == 'bool':
                i['type'] = 'checkbox'
            if i.get('type') == 'float':
                i['type'] = 'numeric'
                i['format'] = '0.00'
            elif i.get('type') == 'int':
                i['type'] = 'numeric'
            return i

        return [map_type(i) for i in l]

    @property
    def settings(self):
        settings = dict(
            columns=self.columns,
            stretchH='all',
            autoWrapRow=True,
            nestedHeaders=self.nested_headers,
            minSpareRows=1,
        )
        return settings

    @property
    def widths(self):
        widths = [self.sheet.column_dimensions[cell.column].width for cell in self.names_cells]
        widths = [int(width) if width else 5.0 for width in widths]
        widths = [int(width) * 4 for width in widths]
        return widths

    @classmethod
    def from_title_cell(cls, sheet, cell, color_names):
        title = cell.value
        slug = sheet.cell(row=cell.row, column=cell.col_idx + 1).value

        min_row = cell.row + 1
        min_col = cell.col_idx

        def get_color(cell):
            try:
                return cell.fill.fgColor.rgb
            except:
                pass

        def check_color(cell):
            color = get_color(cell)
            if color is None or color == '00000000':
                return False
            return True

        def check_continue(c):
            return c.coordinate in sheet.merged_cells or check_color(c)

        def get_name_by_color(color_names, color):
            return color_names.get(color, 'unknown-{}'.format(color))

        cols = sheet.iter_cols(min_col=min_col, min_row=min_row, max_row=min_row)
        cols = [col[0] for col in cols]
        max_col = list(takewhile(check_continue, cols))[-1].col_idx

        rows = sheet.iter_rows(min_row=min_row, min_col=min_col, max_col=max_col)
        rows = [i[0] for i in rows]
        max_row = list(takewhile(check_continue, rows))[-1].row

        rows = sheet.get_squared_range(min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)
        name_cells_rows = {get_name_by_color(color_names, k): list(v) for k, v in
                           groupby(rows, lambda row: get_color(row[0]))}
        return Item(sheet=sheet, title=title, slug=slug, name_cells_rows=name_cells_rows)


# print(make_box(c_titles[0]))





def xlsx_to_configs_samples(path, sheet_name=None, sheet_index=0):
    workbook = load_workbook(path)
    if not sheet_name:
        sheet_name = workbook.get_sheet_names()[sheet_index]
    sheet = workbook.get_sheet_by_name(sheet_name)
    cells = sorted(sheet.get_cell_collection(), key=lambda c: (c.row, c.col_idx))
    name_colors = {c.value: c.fill.fgColor.rgb for c in cells if c.column == 'A'}
    color_names = {y: x for x, y in name_colors.items()}

    def match_color(cell, color):
        try:
            return cell.fill.fgColor.rgb == color
        except:
            return False

    cell_titles = [c for c in cells if c.column != 'A' and match_color(c, name_colors['title'])]
    samples = OrderedDict()
    configs = []

    for cell_title in cell_titles:
        i = Item.from_title_cell(sheet, cell_title, color_names)
        samples[i.slug] = i.samples
        config = dict(
            title=i.title,
            slug=i.slug,
            settings=i.settings,
        )
        configs.append(config)
    return configs, samples


if __name__ == '__main__':
    p = '0.xlsx'
    configs, samples = xlsx_to_configs_samples(p)
    print(yaml_dump(configs))
    print('-' * 20)
    print(yaml_dump(samples))
